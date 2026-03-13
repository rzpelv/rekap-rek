#!/usr/bin/env python3
"""
BRI Rekening Koran → Excel Rekap
Cara pakai:
  python bri_rekap.py file.pdf
  python bri_rekap.py folder_berisi_pdf/
"""

import sys, re, os
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("Install dulu: pip install pdfplumber openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Install dulu: pip install openpyxl")
    sys.exit(1)

MONTHS_ID = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

# ── Keyword yang PASTI bukan penjualan ────────────────────────────────────────
# Prinsip: HANYA yang benar-benar bukan penerimaan dari customer/pihak luar.
# Semua kredit dari pihak luar (RTGS, transfer, dsb.) = Penjualan by default.

NON_PENJ_KW = [
    # Bunga & biaya bank
    "INTEREST ON",
    "BUNGA TABUNGAN",
    "BUNGA DEPOSITO",
    " TAX",
    "BIAYA ADM",
    "BIAYA PROVISI",
    "ADMINISTRASI",
    # Transfer antar rekening sendiri / internal
    "OVERBOOKING",
    "BY SURAT REF",
    "IFT_TO",
    "IFT TO",
    "FROM:0",            # transfer dari nomor rekening sendiri
    "TO:0",
    "PMDH BUKUAN",
    "OB ESCROW",
    "ESCROW",
    "ESB:RTGS:",         # RTGS keluar (outgoing), bukan penerimaan
    # Pencairan kredit / pinjaman dari bank (bukan dari customer)
    "PENCAIRAN KRD",
    "PENCAIRAN",         # semua jenis pencairan (kredit, deposito, dll) = Non penjualan
    "CAIRKAN",
    "CAIR KRD",
    "PMBY PINJ",
    "PLAFON",
    "FASILITAS KRD",
    "KMK",
    "KPR",
    "DROPING",
    "DROPPING",
    "REALISASI KRD",
]

# Whole-word keywords (regex \b...\b) — lebih hati-hati agar tidak false positive
NON_PENJ_WHOLE = [
    r'\bNAS\b',          # kode NAS (bukan NASIONAL, ANAS, dll)
    r'\bBRIVA\b',
]

# Keyword yang diprioritaskan sebagai Penjualan (override default)
PENJ_KW = [
    # Instansi pemerintah / APBD / APBN
    "SP2D", "SPAN", "KASDA", "APBD", "APBN",
    "PEMKOT", "PEMKAB", "PEMPROV", "PEMERINTAH",
    # Fasilitas kesehatan
    "RSUD", "RSIA", "RSUP", "RSU ", "RS ",
    "RUMAH SAKIT", "PUSKESMAS", "KLINIK", "LABORATORIUM",
    # Pendidikan
    "SEKOLAH", "UNIVERSITAS", "UNIV ", "AKADEMI", "MADRASAH",
    "PESANTREN", "SDN ", "SMPN ", "SMAN ",
    # BUMN / perusahaan besar
    "PLN ", "PDAM", "PERTAMINA", "TELKOM",
    "PUPUK KALTIM", "PUPUK INDONESIA", "INKA MULTI",
    # Pola pembayaran
    "PAYMENT PT", "PEMBAYARAN PT", "BANK MANDIRI-PEMBAYARA",
    "BANK BNI-PEMB", "S2P",
    # Kode internal yang diketahui = penjualan
    "ESB:INDS:",
]

# Kata yang menandai akhir blok transaksi (baris summary / footer)
STOP_KW = [
    'Saldo Awal','Opening Balance','Terbilang','In Words',
    'Biaya materai','Revenue Stamp','Apabila terdapat','In the case',
    'Salinan rekening','The copy of','Apabila ada perubahan','Should there be',
    'Created By','BRISIM',
]

# Baris header tabel yang harus dilewati
SKIP_KW = [
    'Tanggal Transaksi','Transaction Date','LAPORAN TRANSAKSI',
    'Halaman','Page ','Printed By','Business Unit',
    'Nama Produk','Product Name','Valuta','Currency',
    'No. Rekening','Account No','Kepada Yth','Transaction Period',
    'Transaction Periode','STATEMENT OF','Statement Date',
    'Tanggal Laporan',
]

NUM_RE    = re.compile(r'^[\d,]+\.\d{2}$')
DATE_RE   = re.compile(r'^\d{2}/\d{2}/\d{2}$')
TIME_RE   = re.compile(r'^\d{2}:\d{2}:\d{2}$')
TELLER_RE = re.compile(r'^\d{7}$')

# ── Warna Excel ────────────────────────────────────────────────────────────────
CLR = {
    "title_bg" : "1F3864",  "title_fg" : "FFFFFF",
    "sub_bg"   : "2E75B6",  "sub_fg"   : "FFFFFF",
    "hdr_bg"   : "2E75B6",  "hdr_fg"   : "FFFFFF",
    "total_bg" : "FFF2CC",
    "alt_bg"   : "F2F7FB",
    "penj_bg"  : "E2EFDA",
    "debet_fg" : "C00000",
    "kredit_fg": "375623",
    "border"   : "BDD7EE",
}
NUM_FMT = '#,##0.00'

def thin_border():
    s = Side(style='thin', color=CLR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def af(clr): return PatternFill("solid", fgColor=clr)

def style_hdr(ws, row_num, bg=None, fg=None):
    bg = bg or CLR["hdr_bg"]; fg = fg or CLR["hdr_fg"]
    for cell in ws[row_num]:
        cell.font      = Font(name='Arial', bold=True, color=fg, size=10)
        cell.fill      = af(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border()

def style_total(ws, row_num):
    for cell in ws[row_num]:
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.fill = af(CLR["total_bg"]); cell.border = thin_border()

def reg(bold=False, color="000000"):
    return Font(name='Arial', bold=bold, color=color, size=10)

# ── Deteksi kolom dari baris header tabel ─────────────────────────────────────
def detect_columns(rows):
    """
    Cari baris header 'Tanggal Transaksi ... Debet ... Kredit ... Saldo'
    dan gunakan posisi X-nya untuk menentukan batas kolom.
    Fallback ke nilai default jika tidak ditemukan.
    """
    for y in sorted(rows.keys()):
        row_words = sorted(rows[y], key=lambda w: w['x0'])
        texts = [w['text'] for w in row_words]
        line  = ' '.join(texts)
        if 'Debet' in line and 'Kredit' in line and 'Saldo' in line:
            col_map = {}
            for w in row_words:
                t = w['text']
                if t == 'Debet':  col_map['debet']  = w['x0']
                if t == 'Kredit': col_map['kredit'] = w['x0']
                if t == 'Saldo':  col_map['saldo']  = w['x0']
                if t in ('Teller','User'): col_map['teller'] = w['x0']
            if 'debet' in col_map:
                # Teller biasanya ~30px sebelum Debet
                teller_x = col_map.get('teller', col_map['debet'] - 30)
                return {
                    'teller': teller_x - 5,
                    'debet' : col_map['debet']  - 5,
                    'kredit': col_map['kredit'] - 5,
                    'saldo' : col_map['saldo']  - 5,
                }
    # Default (format lama)
    return {'teller': 295, 'debet': 360, 'kredit': 465, 'saldo': 560}

# ── Temukan Y baris summary (Saldo Awal) ──────────────────────────────────────
def find_summary_y(rows):
    for y in sorted(rows.keys()):
        line = ' '.join(w['text'] for w in sorted(rows[y], key=lambda w: w['x0']))
        if 'Saldo Awal' in line or 'Opening Balance' in line:
            return y
    return 99999

# ── Parse satu PDF ─────────────────────────────────────────────────────────────
def parse_pdf(pdf_path):
    meta = {
        "accountNo": "", "companyName": "", "period": "",
        "opening": 0, "totalDebet": 0, "totalKredit": 0, "closing": 0
    }
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            rows  = defaultdict(list)
            for w in words:
                rows[round(w['top'] / 2) * 2].append(w)

            # ── Deteksi kolom dari header tabel ──
            cols = detect_columns(rows)

            # ── Temukan Y baris summary ──
            summary_y = find_summary_y(rows)

            # ── Meta dari halaman 1 ──
            if page_num == 0:
                for y in sorted(rows.keys()):
                    rw   = sorted(rows[y], key=lambda w: w['x0'])
                    line = ' '.join(w['text'] for w in rw)

                    # No. Rekening: ambil angka 15 digit setelah ':'
                    if not meta["accountNo"]:
                        m = re.search(r':\s*(\d{15})', line)
                        if m: meta["accountNo"] = m.group(1)

                    # Nama perusahaan: baris yang mengandung nama tapi bukan header
                    # Ambil teks di sisi kiri (x < 300) yang bukan keyword
                    if not meta["companyName"] and y > 100 and y < 200:
                        left_words = [w['text'] for w in rw
                                      if w['x0'] < 300
                                      and w['text'] not in ('Kepada','Yth.','/','To',':')
                                      and not re.match(r'^\d', w['text'])]
                        candidate = ' '.join(left_words).strip()
                        # Harus minimal 3 karakter dan semua huruf besar (nama perusahaan)
                        if (len(candidate) >= 3
                                and candidate.replace(' ','').isupper()
                                and candidate not in ('RAYA','JL','LT','RT')):
                            meta["companyName"] = candidate

                    # Periode transaksi
                    pm = re.search(r'(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})', line)
                    if pm and not meta["period"]:
                        meta["period"] = f"{pm.group(1)} - {pm.group(2)}"

            # ── Summary (Saldo Awal dst.) ──
            for y in sorted(rows.keys()):
                rw   = sorted(rows[y], key=lambda w: w['x0'])
                nums = [w for w in rw if NUM_RE.match(w['text'])]
                if len(nums) >= 4:
                    vs = [float(n['text'].replace(',', '')) for n in nums]
                    if vs[1] > 1e5:  # Total Debet pasti besar
                        meta["opening"]     = vs[0]
                        meta["totalDebet"]  = vs[1]
                        meta["totalKredit"] = vs[2]
                        meta["closing"]     = vs[3]

            # ── Parse transaksi ──
            current_tx = None

            for y in sorted(rows.keys()):
                # STOP: jangan proses di atas / sama dengan baris summary
                if y >= summary_y:
                    break

                row_words = sorted(rows[y], key=lambda w: w['x0'])
                full_text = ' '.join(w['text'] for w in row_words)

                # Skip baris header/info
                if any(s in full_text for s in SKIP_KW):
                    continue
                # Skip baris footer/stop
                if any(s in full_text for s in STOP_KW):
                    break

                first  = row_words[0]['text'] if row_words else ''
                is_tx  = bool(DATE_RE.match(first))

                if is_tx:
                    if current_tx:
                        transactions.append(current_tx)
                    date_str = ''
                    desc_parts = []
                    debet = kredit = balance = 0.0

                    for w in row_words:
                        t = w['text']; x = w['x0']
                        if DATE_RE.match(t) and not date_str:
                            date_str = t
                        elif TIME_RE.match(t) and date_str and ' ' not in date_str:
                            date_str += ' ' + t
                        elif TELLER_RE.match(t):
                            pass  # skip teller ID
                        elif NUM_RE.match(t):
                            v = float(t.replace(',', ''))
                            if   x >= cols['saldo']:   balance = v
                            elif x >= cols['kredit']:  kredit  = v
                            elif x >= cols['debet']:   debet   = v
                            # angka di kiri cols['debet'] = bagian deskripsi (no. rekening, dll)
                        elif x < cols['teller']:
                            # Bagian deskripsi
                            if not DATE_RE.match(t) and not TIME_RE.match(t):
                                desc_parts.append(t)

                    current_tx = {
                        'date': date_str,
                        'desc': ' '.join(desc_parts),
                        'debet': debet, 'kredit': kredit, 'balance': balance
                    }

                elif current_tx:
                    # Baris lanjutan deskripsi — hanya jika tidak ada angka di kolom kanan
                    right_nums = [w for w in row_words
                                  if NUM_RE.match(w['text']) and w['x0'] >= cols['debet']]
                    if not right_nums:
                        extra = ' '.join(
                            w['text'] for w in row_words
                            if w['x0'] < cols['teller']
                            and not TELLER_RE.match(w['text'])
                            and not NUM_RE.match(w['text'])
                        ).strip()
                        if extra:
                            current_tx['desc'] = (current_tx['desc'] + ' ' + extra).strip()

            if current_tx:
                transactions.append(current_tx)
                current_tx = None

    # Tambah bulan & kategori
    for tx in transactions:
        tx['month']    = _month_key(tx['date'])
        tx['kategori'] = _categorize(tx['desc'], meta['companyName'], tx['debet'], tx['kredit'])

    return meta, transactions


def _month_key(d):
    m = re.match(r'^(\d{2})/(\d{2})/(\d{2})', d or '')
    if not m: return 'Unknown'
    return f"{MONTHS_ID[int(m.group(2))-1]} 20{m.group(3)}"

def _contains_own_name(desc, company_name):
    """
    Cek apakah deskripsi mengandung nama perusahaan sendiri (sebagian atau seluruhnya).
    Strategi (dari ketat ke longgar):
    1. Nama lengkap persis
    2. Kombinasi 2 kata berurutan dari nama (minimal 1 kata >= 4 huruf)
    3. Nama terpotong: kata pertama + awal kata berikutnya (misal 'NUGA SIGMA P')
    4. Kata tunggal yang sangat spesifik / unik (>= 8 huruf, bukan kata umum)
    """
    if not company_name:
        return False

    up_desc = (desc or '').upper()
    up_name = company_name.upper()

    # Hapus prefix badan hukum yang tidak spesifik
    PREFIX = {'PT', 'CV', 'UD', 'PD', 'TB', 'TBK', 'TBKK', 'AN', 'THE'}
    all_words = up_name.split()
    name_words = [w for w in all_words if w not in PREFIX]

    # 1. Nama lengkap
    if up_name in up_desc:
        return True

    # 2. Kombinasi 2 kata berurutan (minimal 1 kata >= 4 huruf, total >= 7 karakter)
    for i in range(len(name_words) - 1):
        w1, w2 = name_words[i], name_words[i+1]
        if max(len(w1), len(w2)) >= 4:
            bigram = w1 + ' ' + w2
            if bigram in up_desc:
                return True

    # 3. Nama terpotong: cek "KATA1 KATA2_AWALAN" (nama dipotong di tengah kata)
    #    Contoh: "NUGA SIGMA POTENZIA" -> cek "NUGA SIGMA P", "NUGA SIGMA PO", dst.
    for i in range(len(name_words) - 1):
        w1, w2 = name_words[i], name_words[i+1]
        if len(w1) >= 4:
            for trunc in range(1, len(w2)):  # SIGMA P, SIGMA PO, SIGMA POT, ...
                pattern = w1 + ' ' + w2[:trunc]
                # Pastikan pattern diikuti spasi/non-huruf (bukan bagian kata lain)
                import re
                if re.search(re.escape(pattern) + r'(?:\s|$|[^A-Z])', up_desc):
                    return True

    # 4. Kata tunggal sangat spesifik (>= 8 huruf, bukan kata generik)
    GENERIC = {
        'MANDIRI','BERSAMA','SEJAHTERA','INDONESIA','UTAMA','JAYA','MAKMUR',
        'ABADI','MAJU','SENTOSA','PRIMA','NUSANTARA','PERSADA','SARANA',
        'PRATAMA','PERDANA','SETIA','KARYA','PUTRA','PUTRI','ANDALAN',
        'SUKSES','MULIA','LANCAR','AMANAH','BERKAH','INDAH','AGUNG',
    }
    for w in name_words:
        if len(w) >= 8 and w not in GENERIC and w in up_desc:
            return True

    return False

def _is_company_like(name_part):
    """
    Apakah suatu string terlihat seperti nama badan usaha / instansi
    (bukan nama orang, bukan nomor rekening, bukan kode sistem)?
    """
    # Tanda badan usaha
    ENTITY_MARKERS = ['PT','CV','UD','PD','BANK','DINAS','KEMENTERIAN',
                      'PEMKOT','PEMKAB','PEMPROV','BRI','BNI','MANDIRI',
                      'KOPERASI','YAYASAN','BPJS','KAI','PLN','PERTAMINA',
                      'TELKOM','BULOG','PERUM','PERSEROAN','TBKK','TBK']
    up = name_part.upper()
    for m in ENTITY_MARKERS:
        if m in up:
            return True
    # Lebih dari 2 kata huruf besar semua → kemungkinan nama perusahaan/instansi
    words = [w for w in up.split() if w.isalpha() and len(w) >= 3]
    if len(words) >= 2 and all(w == w.upper() for w in words):
        return True
    return False

# Nama kota / entitas generik yang muncul setelah RTGS# → bukan customer nyata
RTGS_NON_PENJ_NAMES = set()  # tidak dipakai — RTGS dari manapun = Penjualan kecuali nama sendiri

def _make_abbreviations(company_name):
    """
    Buat daftar singkatan dari nama perusahaan.
    Contoh: KARYA PUTRA ANDALAN → KPA, NUGA SIGMA POTENZIA → NSP
    Termasuk kombinasi 2 kata: KP, PA, NS, SP, dll
    """
    PREFIX = {'PT','CV','UD','PD','TB','TBK','AN','THE'}
    words = [w for w in company_name.upper().split() if w not in PREFIX and len(w) >= 2]
    abbrevs = set()
    if not words:
        return abbrevs
    # Singkatan semua kata: KPA, NSP, dll
    abbrevs.add(''.join(w[0] for w in words))
    # Singkatan 2 kata pertama: KP
    if len(words) >= 2:
        abbrevs.add(words[0][0] + words[1][0])
    # Singkatan 2 kata terakhir: PA
    if len(words) >= 2:
        abbrevs.add(words[-2][0] + words[-1][0])
    # Hanya yang >= 2 karakter
    return {a for a in abbrevs if len(a) >= 2}

def _rtgs_is_own_or_generic(desc, company_name):
    """
    RTGS# → Non penjualan HANYA jika nama pengirim mengandung nama rekening sendiri.
    Semua RTGS dari pihak lain = Penjualan.
    Format: RTGS#NAMA_PENGIRIM RTGS STP ESB:... atau RTGS#NAMA #KODE
    """
    import re as _re
    up = desc.upper()
    m = _re.match(r'RTGS#(.+?)(?:\s+(?:PT\s+)?RTGS\s+STP|\s*#|\s+ESB:|\s+\d{10,}|$)', up)
    if not m:
        return False
    sender = m.group(1).strip()
    # Non penjualan hanya jika sender = nama rekening sendiri
    return _contains_own_name(sender, company_name)


def _categorize(desc, company_name='', debet=0, kredit=0):
    """
    Kategorisasi transaksi — filosofi AGRESIF (lebih banyak Penjualan, koreksi manual):

    URUTAN PRIORITAS:
      1. Debet → Non penjualan
      2. Kredit = 0 → Non penjualan
      3. Nama rekening sendiri → Non penjualan
      4. NON_PENJ_WHOLE (regex whole-word) → Non penjualan
      5. NON_PENJ_KW (substring) → Non penjualan
      6. RTGS# kredit → Penjualan kecuali sender = nama sendiri/generik
      7. PENJ_KW → Penjualan (konfirmasi eksplisit)
      8. DEFAULT kredit → Penjualan  ← kunci: semua kredit dari luar = penjualan
         (user bisa koreksi manual via dropdown)
    """
    import re as _re
    up = (desc or '').upper()

    # Rule 1: Debet tidak pernah Penjualan
    if debet > 0 and kredit == 0:
        return 'Non penjualan'

    # Rule 2: Tidak ada kredit masuk
    if kredit == 0:
        return 'Non penjualan'

    # Rule 3: Mengandung nama rekening sendiri → Non penjualan
    if _contains_own_name(desc, company_name):
        return 'Non penjualan'

    # Rule 4: Whole-word keywords non-penjualan
    for pattern in NON_PENJ_WHOLE:
        if _re.search(pattern, up):
            return 'Non penjualan'

    # Rule 5: Substring keywords non-penjualan
    for kw in NON_PENJ_KW:
        if kw.upper() in up:
            return 'Non penjualan'

    # Rule 6: RTGS# kredit — cek apakah sender nama sendiri atau nama generik
    if up.startswith('RTGS#'):
        if _rtgs_is_own_or_generic(desc, company_name):
            return 'Non penjualan'
        return 'Penjualan'   # RTGS dari pihak luar = Penjualan

    # Rule 7: Keyword Penjualan eksplisit (konfirmasi tambahan, sebenarnya sudah default)
    for kw in PENJ_KW:
        if kw.upper() in up:
            return 'Penjualan'

    # Rule 8: DEFAULT — semua kredit dari pihak luar = Penjualan
    # User bisa koreksi via dropdown jika salah
    return 'Penjualan'


# ── Buat Excel ─────────────────────────────────────────────────────────────────
def build_excel(all_transactions, meta, out_path):
    from openpyxl.worksheet.datavalidation import DataValidation

    by_month = defaultdict(list)
    for tx in all_transactions:
        by_month[tx['month']].append(tx)

    month_order = [f"{m} {y}" for y in range(2024, 2027)
                   for m in MONTHS_ID if f"{m} {y}" in by_month]

    n_tx = len(all_transactions)   # total baris di Edit Penjualan
    EP   = "'Edit Penjualan'"       # nama sheet untuk formula cross-ref

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Edit Penjualan  (MASTER — user edit di sini)
    # Kolom: A=No | B=Periode | C=Tanggal | D=Deskripsi | E=Debet | F=Kredit | G=Kategori
    # ═══════════════════════════════════════════════════════════════════════════
    ws_ep = wb.active
    ws_ep.title = "Edit Penjualan"

    # Title
    ws_ep.merge_cells('A1:G1')
    ws_ep['A1'] = f"✏️  EDIT KATEGORI PENJUALAN  —  {meta['companyName']}  —  {meta['accountNo']}"
    ws_ep['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=12)
    ws_ep['A1'].fill      = af(CLR["title_bg"])
    ws_ep['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ep.row_dimensions[1].height = 26

    ws_ep.merge_cells('A2:G2')
    ws_ep['A2'] = "👆  Ubah kolom KATEGORI (G) untuk menandai transaksi sebagai Penjualan atau Non penjualan — Summary otomatis terupdate"
    ws_ep['A2'].font      = Font(name='Arial', italic=True, color="7F7F7F", size=9)
    ws_ep['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ep.row_dimensions[2].height = 16

    hdrs_ep = ["No","Periode","Posting Date","Transaction Description","Debet (Rp)","Kredit (Rp)","Kategori"]
    ws_ep.append(hdrs_ep)
    style_hdr(ws_ep, 3)
    ws_ep.row_dimensions[3].height = 26

    # Dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"Penjualan,Non penjualan"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Nilai tidak valid",
        error='Pilih: "Penjualan" atau "Non penjualan"'
    )
    ws_ep.add_data_validation(dv)

    # Isi data
    for i, tx in enumerate(all_transactions):
        r = i + 4   # baris 4 dst
        penj = tx['kategori'] == 'Penjualan'
        ws_ep.append([
            i + 1,
            tx['month'],
            tx['date'],
            tx['desc'],
            tx['debet']  if tx['debet']  else None,
            tx['kredit'] if tx['kredit'] else None,
            tx['kategori']
        ])
        bg = CLR["penj_bg"] if penj else (CLR["alt_bg"] if i % 2 == 0 else "FFFFFF")
        for c in range(1, 8):
            cell = ws_ep.cell(row=r, column=c)
            cell.fill   = af(bg)
            cell.border = thin_border()
            cell.font   = reg()
        ws_ep.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws_ep.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws_ep.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        e = ws_ep.cell(row=r, column=5)
        f = ws_ep.cell(row=r, column=6)
        e.number_format = NUM_FMT; e.alignment = Alignment(horizontal='right')
        f.number_format = NUM_FMT; f.alignment = Alignment(horizontal='right')
        if tx['debet']:  e.font = reg(color=CLR["debet_fg"])
        if tx['kredit']: f.font = reg(color=CLR["kredit_fg"])
        kat_cell = ws_ep.cell(row=r, column=7)
        kat_cell.alignment = Alignment(horizontal='center')
        if penj: kat_cell.font = reg(bold=True, color=CLR["kredit_fg"])
        dv.add(kat_cell)

    for col, w in zip("ABCDEFG", [6, 14, 20, 62, 18, 18, 16]):
        ws_ep.column_dimensions[col].width = w
    ws_ep.freeze_panes = 'A4'
    ws_ep.auto_filter.ref = f"A3:G{n_tx + 3}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Summary  (formula SUMIFS ke Edit Penjualan)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Summary")

    ws.merge_cells('A1:H1')
    ws['A1'] = "REKAP REKENING KORAN BRI"
    ws['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=13)
    ws['A1'].fill      = af(CLR["title_bg"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:H2')
    ws['A2'] = f"No. Rek: {meta['accountNo']}   |   {meta['companyName']}   |   Periode: {meta['period']}"
    ws['A2'].font      = Font(name='Arial', bold=True, color=CLR["sub_fg"], size=10)
    ws['A2'].fill      = af(CLR["sub_bg"])
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.append([])  # baris 3 kosong

    hdrs = ["Periode","Jumlah Transaksi","Total Debet (Rp)","Total Kredit (Rp)",
            "Total Penjualan (Rp)","Saldo Awal (Rp)","Saldo Akhir (Rp)","Status"]
    ws.append(hdrs)
    style_hdr(ws, 4)
    ws.row_dimensions[4].height = 30

    # Range Edit Penjualan untuk formula (baris 4..n_tx+3)
    ep_b = f"{EP}!$B$4:$B${n_tx+3}"   # kolom Periode
    ep_e = f"{EP}!$E$4:$E${n_tx+3}"   # kolom Debet
    ep_f = f"{EP}!$F$4:$F${n_tx+3}"   # kolom Kredit
    ep_g = f"{EP}!$G$4:$G${n_tx+3}"   # kolom Kategori

    for i, m in enumerate(month_order):
        txs_m = by_month[m]
        r     = 5 + i
        sa    = txs_m[0]['balance'] - txs_m[0]['kredit'] + txs_m[0]['debet'] if txs_m else 0
        se    = txs_m[-1]['balance'] if txs_m else 0

        # Formula SUMIFS ke Edit Penjualan — otomatis update saat user edit Kategori
        f_jumlah  = f'=COUNTIF({ep_b},A{r})'
        f_debet   = f'=SUMIF({ep_b},A{r},{ep_e})'
        f_kredit  = f'=SUMIF({ep_b},A{r},{ep_f})'
        f_penj    = f'=SUMIFS({ep_f},{ep_b},A{r},{ep_g},"Penjualan")'

        ws.append([m, f_jumlah, f_debet, f_kredit, f_penj, sa, se, "OK"])

        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = reg(); cell.fill = af(bg); cell.border = thin_border()
        for c in [3, 4, 5, 6, 7]:
            ws.cell(row=r, column=c).number_format = NUM_FMT
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=8).font      = reg(bold=True, color=CLR["kredit_fg"])
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    # Baris TOTAL — formula SUM dari baris bulan
    tr       = 5 + len(month_order)
    r_start  = 5
    r_end    = tr - 1
    ws.append([
        "TOTAL",
        f"=SUM(B{r_start}:B{r_end})",
        f"=SUM(C{r_start}:C{r_end})",
        f"=SUM(D{r_start}:D{r_end})",
        f"=SUM(E{r_start}:E{r_end})",
        "", "", ""
    ])
    style_total(ws, tr)
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal='center')
    for c in [3, 4, 5]:
        ws.cell(row=tr, column=c).number_format = NUM_FMT
        ws.cell(row=tr, column=c).alignment = Alignment(horizontal='right')

    # Ringkasan dari PDF
    tr += 2
    ws.cell(row=tr, column=1).value = "── RINGKASAN DARI PDF ──"
    ws.cell(row=tr, column=1).font  = reg(bold=True, color="2E75B6")
    for label, val in [("Saldo Awal", meta['opening']),
                       ("Total Debet", meta['totalDebet']),
                       ("Total Kredit", meta['totalKredit']),
                       ("Saldo Akhir", meta['closing'])]:
        tr += 1
        ws.cell(row=tr, column=1).value = label
        ws.cell(row=tr, column=1).font  = reg()
        ws.cell(row=tr, column=3).value = val
        ws.cell(row=tr, column=3).number_format = NUM_FMT
        ws.cell(row=tr, column=3).alignment = Alignment(horizontal='right')

    for col, w in zip("ABCDEFGH", [16, 18, 22, 22, 22, 22, 22, 10]):
        ws.column_dimensions[col].width = w

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET per Bulan  (Kategori = INDEX ke Edit Penjualan)
    # ═══════════════════════════════════════════════════════════════════════════
    # Buat index global: nomor baris di Edit Penjualan untuk tiap transaksi
    tx_ep_row = {id(tx): (idx + 4) for idx, tx in enumerate(all_transactions)}

    for m in month_order:
        txs_m = by_month[m]
        safe  = m.replace(' ', '_')[:31]
        ws2   = wb.create_sheet(title=safe)

        ws2.merge_cells('A1:G1')
        ws2['A1'] = f"TRANSAKSI {m.upper()}  —  {meta['companyName']}  —  No. Rek: {meta['accountNo']}"
        ws2['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
        ws2['A1'].fill      = af(CLR["title_bg"])
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 22

        ws2.append(["No","Posting Date","Transaction Description",
                    "Debet (Rp)","Kredit (Rp)","Balance (Rp)","Kategori"])
        style_hdr(ws2, 2)
        ws2.row_dimensions[2].height = 26

        td_tot = tk_tot = 0
        for i, tx in enumerate(txs_m):
            ep_row = tx_ep_row[id(tx)]
            r      = i + 3
            penj   = tx['kategori'] == 'Penjualan'

            ws2.append([
                i + 1, tx['date'], tx['desc'],
                tx['debet']  if tx['debet']  else None,
                tx['kredit'] if tx['kredit'] else None,
                tx['balance'],
                f"={EP}!G{ep_row}"   # ← referensi langsung ke Edit Penjualan
            ])

            bg = CLR["penj_bg"] if penj else (CLR["alt_bg"] if i % 2 == 0 else "FFFFFF")
            for c in range(1, 8):
                cell = ws2.cell(row=r, column=c)
                cell.fill = af(bg); cell.border = thin_border(); cell.font = reg()
            ws2.cell(row=r, column=1).alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal='center')
            d = ws2.cell(row=r, column=4)
            k = ws2.cell(row=r, column=5)
            b = ws2.cell(row=r, column=6)
            d.number_format = NUM_FMT; d.alignment = Alignment(horizontal='right')
            k.number_format = NUM_FMT; k.alignment = Alignment(horizontal='right')
            b.number_format = NUM_FMT; b.alignment = Alignment(horizontal='right')
            if tx['debet']:  d.font = reg(color=CLR["debet_fg"])
            if tx['kredit']: k.font = reg(color=CLR["kredit_fg"])
            ws2.cell(row=r, column=7).alignment = Alignment(horizontal='center')
            if penj: ws2.cell(row=r, column=7).font = reg(bold=True, color=CLR["kredit_fg"])
            td_tot += tx['debet']; tk_tot += tx['kredit']

        tr2 = len(txs_m) + 3
        ws2.append(["", "TOTAL", "", td_tot, tk_tot, "", ""])
        style_total(ws2, tr2)
        ws2.cell(row=tr2, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=tr2, column=4).number_format = NUM_FMT
        ws2.cell(row=tr2, column=4).alignment = Alignment(horizontal='right')
        ws2.cell(row=tr2, column=5).number_format = NUM_FMT
        ws2.cell(row=tr2, column=5).alignment = Alignment(horizontal='right')

        for col, w in zip("ABCDEFG", [6, 20, 58, 20, 20, 20, 16]):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = 'A3'
        ws2.auto_filter.ref = f"A2:G{tr2}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET Penjualan Customer  (data statis — snapshot saat export)
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Penjualan Customer")
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "PENJUALAN CUSTOMER"
    ws3['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
    ws3['A1'].fill      = af(CLR["title_bg"])
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 22
    ws3.append(["Periode","Posting Date","Transaction Description","Kredit (Rp)","Customer"])
    style_hdr(ws3, 2)

    penj_txs = [tx for tx in all_transactions if tx['kategori'] == 'Penjualan']
    for i, tx in enumerate(penj_txs):
        ep_row = tx_ep_row[id(tx)]
        r      = i + 3
        ws3.append([
            tx['month'], tx['date'], tx['desc'],
            tx['kredit'],
            f"={EP}!D{ep_row}"   # deskripsi dari Edit Penjualan
        ])
        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c)
            cell.fill = af(bg); cell.border = thin_border(); cell.font = reg()
        ws3.cell(row=r, column=4).number_format = NUM_FMT
        ws3.cell(row=r, column=4).alignment     = Alignment(horizontal='right')
        ws3.cell(row=r, column=4).font          = reg(color=CLR["kredit_fg"])

    for col, w in zip("ABCDE", [14, 20, 58, 20, 50]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = 'A3'

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET Customer Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Customer Summary")
    ws4.merge_cells('A1:C1')
    ws4['A1'] = "RINGKASAN CUSTOMER PENJUALAN"
    ws4['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
    ws4['A1'].fill      = af(CLR["title_bg"])
    ws4['A1'].alignment = Alignment(horizontal='center')
    ws4.row_dimensions[1].height = 22
    ws4.append(["Periode","Customer","Kredit (Rp)"])
    style_hdr(ws4, 2)

    cs_map = {}
    for tx in penj_txs:
        key = f"{tx['month']}||{tx['desc'][:60]}"
        if key not in cs_map:
            cs_map[key] = {'periode': tx['month'], 'customer': tx['desc'][:80], 'kredit': 0}
        cs_map[key]['kredit'] += tx['kredit']

    for i, rv in enumerate(sorted(cs_map.values(), key=lambda x: -x['kredit'])):
        ws4.append([rv['periode'], rv['customer'], rv['kredit']])
        rn = i + 3
        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"
        for c in range(1, 4):
            cell = ws4.cell(row=rn, column=c)
            cell.fill = af(bg); cell.border = thin_border(); cell.font = reg()
        ws4.cell(row=rn, column=3).number_format = NUM_FMT
        ws4.cell(row=rn, column=3).alignment     = Alignment(horizontal='right')
        ws4.cell(row=rn, column=3).font          = reg(color=CLR["kredit_fg"])

    for col, w in zip("ABC", [14, 68, 22]):
        ws4.column_dimensions[col].width = w

    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("  BRI Rekening Koran → Excel Rekap")
        print("=" * 60)
        print("\nCara pakai:")
        print("  python bri_rekap.py file.pdf")
        print("  python bri_rekap.py folder/")
        sys.exit(0)

    input_path = Path(sys.argv[1])

    if input_path.is_dir():
        pdf_files = sorted(input_path.glob("*.pdf"))
        if not pdf_files:
            print(f"Tidak ada file PDF di folder: {input_path}"); sys.exit(1)
        print(f"Ditemukan {len(pdf_files)} file PDF")
    elif input_path.is_file():
        pdf_files = [input_path]
    else:
        print(f"File/folder tidak ditemukan: {input_path}"); sys.exit(1)

    all_transactions = []
    meta_combined    = {"accountNo":"","companyName":"","period":"",
                        "opening":0,"totalDebet":0,"totalKredit":0,"closing":0}
    periods = []

    for pdf_file in pdf_files:
        print(f"\nMemproses: {pdf_file.name}")
        meta, txs = parse_pdf(pdf_file)
        print(f"  Rekening  : {meta['accountNo']} — {meta['companyName']}")
        print(f"  Periode   : {meta['period']}")
        print(f"  Transaksi : {len(txs)}")
        all_transactions.extend(txs)
        if not meta_combined["accountNo"] and meta["accountNo"]:
            meta_combined["accountNo"]   = meta["accountNo"]
            meta_combined["companyName"] = meta["companyName"]
        meta_combined["totalDebet"]  += meta["totalDebet"]
        meta_combined["totalKredit"] += meta["totalKredit"]
        if meta["period"]: periods.append(meta["period"])
        if not meta_combined["opening"] and meta["opening"]:
            meta_combined["opening"] = meta["opening"]
        if meta["closing"]:
            meta_combined["closing"] = meta["closing"]

    if periods:
        meta_combined["period"] = f"{periods[0].split(' - ')[0]} - {periods[-1].split(' - ')[-1]}"

    acc = meta_combined["accountNo"] or "rekening"
    if input_path.is_file():
        out = Path.cwd() / f"rekap_{input_path.stem}.xlsx"
    else:
        out = Path.cwd() / f"rekap_bri_{acc}.xlsx"

    print(f"\nMembuat Excel...")
    build_excel(all_transactions, meta_combined, str(out))

    print(f"\n✓ SELESAI! File tersimpan di:")
    print(f"  {out.resolve()}")
    print(f"\n  Total transaksi  : {len(all_transactions)}")
    penj = [t for t in all_transactions if t['kategori'] == 'Penjualan']
    print(f"  Penjualan        : {len(penj)} transaksi")
    months = sorted(set(t['month'] for t in all_transactions))
    print(f"  Periode          : {', '.join(months)}")


if __name__ == "__main__":
    main()
